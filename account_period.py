# -*- coding: utf-8 -*-

import sys
import os
import time
import openpyxl
import tqdm
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from strings import *


class AccountPeriod:
    def __init__(self, filename):
        self.execute_path = sys.path[0]
        self.filename = filename.split(os.sep)[-1]
        self.read_path = filename if STRING_INPUT in filename else "{0}{1}{2}{3}{4}".format(
            self.execute_path, os.sep, STRING_INPUT, os.sep, filename)
        self.month_period = 24
        self.threshold = 0.9
        self.months = [
            "2019-01", "2019-02", "2019-03", "2019-04", "2019-05", "2019-06",
            "2019-07", "2019-08", "2019-09", "2019-10", "2019-11", "2019-12",
            "2020-01", "2020-02", "2020-03", "2020-04", "2020-05", "2020-06"
            # "2020-07", "2020-08", "2020-09", "2020-10", "2020-11", "2020-12"
        ]
        self.df_input = pd.read_excel(self.read_path, dtype={
            STRING_COL_CUSTOM_NO: str, STRING_COL_REC_BILL_MONEY_SUM: float, STRING_COL_VERIFY_DETAIL_MONEY_SUM: float})
        self.df_custom = self.df_input[[STRING_COL_CUSTOM_NAME, STRING_COL_CUSTOM_NO]].drop_duplicates()
        self.df_input_length = len(self.df_input)
        self.df_input_unit_length = self.df_input_length / self.month_period
        self.hospitals_dictionary = dict(zip(self.df_custom[STRING_COL_CUSTOM_NAME], self.df_custom[STRING_COL_CUSTOM_NO]))
        self.hospitals = sorted(list(self.df_custom[STRING_COL_CUSTOM_NAME]), key=lambda x: x)
        self.data_dictionary = dict()
        
        self.df = pd.DataFrame(columns=LIST_COL_NAMES_DETAILS)
        self.df_summary = pd.DataFrame(columns=LIST_COL_NAMES_DETAILS)
        self.load_data()
        self.build_data()
        self.save_data()
        
    def load_data(self):
        for hospital in self.hospitals:
            child_child_dict = dict()
            child_child_dict["flag"] = False
            child_child_dict["rec"] = 0.0
            child_child_dict["val"] = [0.0] * self.month_period
            child_child_dict["sum"] = [0.0] * self.month_period
            child_dict = dict()
            for one_month in self.months:
                child_dict[one_month] = child_child_dict
            self.data_dictionary[hospital] = child_dict
        hospital = month = ""
        money_list = money_sum = []
        tmp_sum = 0.0
        child_child_dict = dict()
        items = [item for item in self.df_input.iterrows()]
        for index, row in tqdm.tqdm(items, desc="[Step 1/3] Loading"):
            if index % self.month_period == 0:
                hospital = row.get(STRING_COL_CUSTOM_NAME)
                month = row.get(STRING_COL_REC_BILL_MONTH)
                child_child_dict = dict()
                child_child_dict["flag"] = True
                child_child_dict["rec"] = float(row.get(STRING_COL_REC_BILL_MONEY_SUM))
                money_list = []
                money_sum = []
                tmp_sum = 0.0
            tmp = float(row.get(STRING_COL_VERIFY_DETAIL_MONEY_SUM))
            tmp_sum += tmp
            money_list.append(tmp)
            money_sum.append(tmp_sum)
            if (index + 1) % self.month_period == 0:
                child_child_dict["val"] = money_list
                child_child_dict["sum"] = money_sum
                self.data_dictionary[hospital][month] = child_child_dict
    
    def build_data(self):
        for hos_id, hospital in enumerate(tqdm.tqdm(self.hospitals, desc="[Step 2/3] Building")):
            sum_rec = 0.0
            sums = [0.0] * self.month_period
            for m, month in enumerate(self.months):
                line = [hos_id + 1, hospital, self.hospitals_dictionary[hospital], month, self.data_dictionary[hospital][month]["rec"]]
                sum_rec += self.data_dictionary[hospital][month]["rec"]
                for i in range(0, self.month_period):
                    if not self.data_dictionary[hospital][month]["flag"] or self.data_dictionary[hospital][month]["rec"] == 0:
                        line.append(0.0)
                    else:
                        line.append(self.data_dictionary[hospital][month]["sum"][i] / float(self.data_dictionary[hospital][month]["rec"]))
                        sums[i] += self.data_dictionary[hospital][month]["sum"][i]
                self.df = self.df.append(pd.DataFrame({x: [y] for x, y in zip(LIST_COL_NAMES_DETAILS, line)}), ignore_index=True)
            final_line = [hos_id + 1, hospital, self.hospitals_dictionary[hospital], "合计", sum_rec]
            for i in range(0, self.month_period):
                if sum_rec == 0.0:
                    final_line.append(0.0)
                else:
                    final_line.append(self.not_exceed(sums[i] / float(sum_rec)))
            df_final_line = pd.DataFrame({x: [y] for x, y in zip(LIST_COL_NAMES_DETAILS, final_line)})
            self.df = self.df.append(df_final_line, ignore_index=True)
            self.df_summary = self.df_summary.append(df_final_line, ignore_index=True)
        self.df.index += 1
        df_summary_col_name = "达到阈值({0})月数".format(self.threshold)
        self.df_summary[df_summary_col_name] = list((self.df_summary[LIST_COL_NAMES_DETAILS_MONTHS] >= self.threshold).astype(int).sum(axis=1))
        self.df_summary.sort_values(by=[df_summary_col_name, "前23月"], ascending=False, inplace=True)
        self.df_summary.reset_index(drop=True, inplace=True)
        self.df_summary.index += 1
        self.df_summary["序号"] = self.df_summary.index
        list_col_names_details_summary = self.df_summary.columns.values.tolist()[:-1]
        list_col_names_details_summary[3] = df_summary_col_name
        self.df_summary = self.df_summary[list_col_names_details_summary]

    def save_data(self):
        save_path = "{0}{1}output{2}{3}_result_{4}.xlsx".format(
            self.execute_path, os.sep, os.sep, self.filename.split(".")[0], time.strftime("%Y%m%d_%H%M%S", time.localtime(time.time())))
        writer = pd.ExcelWriter(save_path)
        self.df_summary.to_excel(writer, sheet_name="合计汇总", float_format="%.6f", index=False)
        self.df.to_excel(writer, sheet_name="全量应收数据", float_format="%.6f", index=False)
        self.df_input.to_excel(writer, sheet_name="原始HUE导出数据", float_format="%.6f", index=False)
        writer.save()
        self.set_excel_style(save_path, self.threshold)
    
    @staticmethod
    def not_exceed(val, threshold=1.0):
        if val > threshold:
            return threshold
        return val
    
    @staticmethod
    def set_excel_style(filename, threshold):
        book = openpyxl.load_workbook(filename)
        # 初始化字体样式
        font_bold = Font(name="等线",
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline="none",
                         strike=False,
                         color="FF000000",
                         outline="None")
        font_not_bold = Font(name="等线",
                             size=11,
                             italic=False,
                             vertAlign=None,
                             underline="none",
                             strike=False,
                             color="FF000000",
                             outline="None")
        border_full = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"),
                             left=Side(border_style="thin"), right=Side(border_style="thin"))
        for one_sheet_name in tqdm.tqdm(book.sheetnames, desc="[Step 3/3] Beautifying"):
            sheet = book[one_sheet_name]
            max_row = sheet.max_row
            max_column = sheet.max_column
            for row in sheet.rows:
                for cell in row:
                    cell.font = font_not_bold
                    cell.border = border_full
            # 加粗
            for i in range(1, sheet.max_column + 1):
                sheet.cell(1, i).font = font_bold
            # 居中
            for i in range(max_row):
                for j in range(max_column):
                    if i == 0:
                        sheet.cell(i + 1, j + 1).fill = PatternFill(fill_type="solid", fgColor="B4C6E7")
                    sheet.cell(i + 1, j + 1).alignment = Alignment(horizontal="center", vertical="center")
                    if one_sheet_name == "合计汇总" and i > 0 and j >= 5:
                        if sheet.cell(i + 1, j + 1).value >= threshold:
                            sheet.cell(i + 1, j + 1).fill = PatternFill(fill_type="solid", fgColor="FFB5B5")
                        else:
                            sheet.cell(i + 1, j + 1).fill = PatternFill(fill_type="solid", fgColor="C4E1FF")
            max_column_dict = {}
            for i in range(1, max_column + 1):
                for j in range(1, max_row + 1):
                    column = 0
                    sheet_value = sheet.cell(j, i).value
                    sheet_value_list = [k for k in str(sheet_value)]
                    for v in sheet_value_list:
                        if ord(v) < 256:
                            column += 1.2
                        else:
                            column += 2.1
                    try:
                        if not max_column_dict.get(i) or column > max_column_dict[i]:
                            max_column_dict[i] = column
                    except Exception as e:
                        print("Error:", e)
                        max_column_dict[i] = column
            for key, value in max_column_dict.items():
                sheet.column_dimensions[get_column_letter(key)].width = value
            book.save(filename)
    
    @staticmethod
    def print_df(df):
        pd.set_option("display.max_rows", 500)
        pd.set_option("display.max_columns", 500)
        pd.set_option("display.width", 1000)
        print(df)


if __name__ == "__main__":
    # AccountPeriod("1.xlsx")
    if len(sys.argv) < 2:
        print("Please input an existing file in dir 'input'! (like: python account_period.py 1.xlsx)")
    else:
        AccountPeriod(sys.argv[1])
    
    

